#!/usr/local/bin/python3

import openpyxl
import numpy as np
import re


# TODO:
#  1) Build

   # hash_loans = {
   #   1819: {
   #    currency_number: 643,
   #    date: 20160101
   #   }
   # }
# 2) Build
   # Read ./for_acra.xlsx
   # Build
   # hash_row_number = {
   #   1819: 12
   # }


def build_loans_data():
  hash_loans = {}
  book = openpyxl.load_workbook('./form_303.xlsx')
  # ws = wb.create_sheet(title='data')
  sheet = book['Sheet']

  for i in range(3,sheet.max_row+1):
    loan_id = sheet.cell(row=i, column=8).value
    currency = sheet.cell(row=i, column=9).value
    loan_date = sheet.cell(row=i, column=7).value
    init_maturity_date = sheet.cell(row=i, column=3).value
    maturity_date = sheet.cell(row=i, column=4).value
    init_rate = sheet.cell(row=i, column=5).value
    rate = sheet.cell(row=i, column=6).value
    interest_period =  sheet.cell(row=i, column=15).value
    amount =  sheet.cell(row=i, column=10).value
    debt_service = sheet.cell(row=i, column=14).value
    reserve = sheet.cell(row=i, column=13).value

    if loan_id is not None:
      hash_loans[loan_id] = {
        'currency': currency,
        'loan_date': loan_date,
        'init_maturity_date': init_maturity_date,
        'maturity_date': maturity_date,
        'init_rate': init_rate/100,
        'rate': rate/100,
        'interest_period': interest_period,
        'amount': amount/1000,
        'debt_service': debt_service,
        'reserve': reserve/100
      }
  # print(hash_loans)
  return hash_loans


def currency_map():
  hash_currency = {
    643: 'RUR',
    840: 'USD',
    978: 'EUR',
    398: 'Иная'
   }
  # print(hash_currency)
  return hash_currency


def interest_period_map():
  hash_interest_period = {
    1: 'ежемесячно',
    2: 'ежеквартально',
    3: 'раз в полгода',
    4: 'ежегодно',
    5: 'иное',
    7: 'иное'
   }
  # print(hash_interest_period)
  return hash_interest_period

def debt_service_map():
  hash_debt_service = {
    1: 'хорошее',
    2: 'среднее',
    3: 'неудовлетворительное'
   }
  # print(hash_debt_service)
  return hash_debt_service



def read_acra(filename):
  book = openpyxl.load_workbook(filename)
  sheet = book['18_TOP-30_LOANS']
  hash_rows = {}
  for i in range(8,sheet.max_row+1):
    loan_id = sheet.cell(row=i, column=1).value
    if loan_id is not None:
      hash_rows.update({loan_id : i})
  # print(hash_rows)
  return hash_rows


def write_xlsx(filename, hash_loans, hash_rows, hash_currency, hash_interest_period, hash_debt_service):
  book = openpyxl.load_workbook(filename)
  sheet = book['18_TOP-30_LOANS']
  for i in range(8,sheet.max_row+1):
    loan_id = sheet.cell(row=i, column=1).value
    if loan_id in hash_rows:
      sheet.cell(row=i, column=12).value = hash_currency.get(hash_loans.get(loan_id)['currency'])
      sheet.cell(row=i, column=13).value = hash_currency.get(hash_loans.get(loan_id)['currency'])
      sheet.cell(row=i, column=14).value = hash_loans.get(loan_id)['loan_date']
      sheet.cell(row=i, column=15).value = hash_loans.get(loan_id)['init_maturity_date']
      sheet.cell(row=i, column=16).value = hash_loans.get(loan_id)['maturity_date']
      sheet.cell(row=i, column=17).value = hash_loans.get(loan_id)['init_rate']
      sheet.cell(row=i, column=18).value = hash_loans.get(loan_id)['rate']
      sheet.cell(row=i, column=19).value = hash_interest_period.get(hash_loans.get(loan_id)['interest_period'])
      sheet.cell(row=i, column=20).value = hash_loans.get(loan_id)['amount']
      sheet.cell(row=i, column=32).value = hash_debt_service.get(hash_loans.get(loan_id)['debt_service'])
      sheet.cell(row=i, column=37).value = hash_loans.get(loan_id)['reserve']





  # for row in range(2, len(list_rows)):
  #     full_id = list_rows[row]['customer_id']

  #     ws.cell(row=row, column=1).value = list_rows[row]['LLP_group']
  #     ws.cell(row=row, column=2).value = list_rows[row]['customer_id']
  #     ws.cell(row=row, column=3).value = list_rows[row]['customer_name']

  #     if full_id is not None:
  #       ws.cell(row=row, column=4).value = debt_dict.get(str(full_id), 0)
  #       ws.cell(row=row, column=5).value = proc_dict.get(str(full_id), [0,0])[0]
  #       ws.cell(row=row, column=6).value = comis_dict.get(str(full_id), [0,0])[0]
  #       ws.cell(row=row, column=7).value = list_rows[row]['LLP_ratio']
  #       ws.cell(row=row, column=8).value = ws.cell(row=row, column=4).value * ws.cell(row=row, column=7).value
  #       ws.cell(row=row, column=9).value = reserve_dict.get(str(full_id), 0)
  #       ws.cell(row=row, column=10).value = proc_dict.get(str(full_id), [0,0])[1]
  #       ws.cell(row=row, column = 11).value = comis_dict.get(str(full_id), [0,0])[1]

  # ws.cell(row=1, column=1).value = 'LLP_group'
  # ws.cell(row=1, column=2).value = 'customer_id'
  # ws.cell(row=1, column=3).value = 'customer_name'
  # ws.cell(row=1, column=4).value = 'principal_amount'
  # ws.cell(row=1, column=5).value = 'interest_amount'
  # ws.cell(row=1, column=6).value = 'commission_amount'
  # ws.cell(row=1, column=7).value ='LLP_ratio'
  # ws.cell(row=1, column=8).value ='LLP_potential'
  # ws.cell(row=1, column=9).value ='LLP_real'
  # ws.cell(row=1, column=10).value ='LLP_interest_real'
  # ws.cell(row=1, column=11).value ='LLP_commission_real'

  book.save(filename=filename)




if __name__ == "__main__":



 write_xlsx('./for_acra.xlsx', build_loans_data(), read_acra('./for_acra.xlsx'), currency_map(), interest_period_map(), debt_service_map())




#def principal_debt (file_path, )


# Записать каждый блок в excel
#   Пойщи как в цикле пройтись по массиву
#   И записать в нужные клеточки каждую строку
#   Найти сoncat функцию, склеивается много массивов в один
#   Писать сразу в файл