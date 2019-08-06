#!/usr/local/bin/python3

import openpyxl
import numpy as np
import re


# TODO:
#  1) Build

   # hash_loans = {
   #   1819: {
   #    currency_number: 643,
   #    date: 20160101
   #   }
   # }
# 2) Build
   # Read ./for_acra.xlsx
   # Build
   # hash_row_number = {
   #   1819: 12
   # }


def build_loans_data():
  hash_loans = {}
  book = openpyxl.load_workbook('./form_0717303.xlsx')
  # ws = wb.create_sheet(title='data')
  sheet = book['Sheet']

  for i in range(3,sheet.max_row+1):
    loan_id = sheet.cell(row=i, column=3).value
    amount =  sheet.cell(row=i, column=4).value

    if loan_id is not None:
      if amount is not None:
        hash_loans[loan_id] = {
          'amount' : amount/1000
        }
  # print(hash_loans)
  return hash_loans


def read_acra(filename):
  book = openpyxl.load_workbook(filename)
  sheet = book['18_TOP-30_LOANS']
  hash_rows = {}
  for i in range(8,sheet.max_row+1):
    loan_id = sheet.cell(row=i, column=1).value
    if loan_id is not None:
      hash_rows.update({loan_id : i})
  # print(hash_rows)
  return hash_rows


def write_xlsx(filename, hash_loans, hash_rows):
  book = openpyxl.load_workbook(filename)
  sheet = book['18_TOP-30_LOANS']
  for i in range(8,sheet.max_row+1):
    credit_id = sheet.cell(row=i, column=1).value
    if credit_id is not None:
      # print(credit_id)
      if credit_id in hash_loans:
        sheet.cell(row=i, column=22).value = hash_loans.get(credit_id)['amount']
      # print(sheet.cell(row=i, column=21).value)
  book.save(filename=filename)




if __name__ == "__main__":



 write_xlsx('./for_acra.xlsx', build_loans_data(), read_acra('./for_acra.xlsx'))




#def principal_debt (file_path, )


# Записать каждый блок в excel
#   Пойщи как в цикле пройтись по массиву
#   И записать в нужные клеточки каждую строку
#   Найти сoncat функцию, склеивается много массивов в один
#   Писать сразу в файл