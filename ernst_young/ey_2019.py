#!/usr/local/bin/python3

import openpyxl

def account_number(file_path):
  book = openpyxl.load_workbook(file_path)
  sheet = book['01.07.2019']

  for i in range(6,sheet.max_row+1):
    customer_id = sheet.cell(row=i, column=13).value
    account = sheet.cell(row=i, column=67).value

    if customer_id is None or account is None:
      continue
    customer_number = customer_id.split('.')[1]
    account = account.replace(account[-7:], customer_number)
    sheet.cell(row=i, column=67).value =  account
    print(account)

  book.save(filename='303_form.xlsx')













if __name__ == "__main__":
  account_number('./303_form.xlsx')