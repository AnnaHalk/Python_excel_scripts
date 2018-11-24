#!/usr/local/bin/python3

import openpyxl
import none_value as nv
import numpy as np
import re
from openpyxl.utils import get_column_letter
from datetime import date
from datetime import timedelta


def filter_by_LLP_group(sheet, required_LLP_group):
  point_in_xlsx ='\x97'

  res = [];
  added_customer_id = {}

  for i in range(12,sheet.max_row+1):

    customer_id = sheet.cell(row=i, column=1).value
    customer_name = sheet.cell(row=i, column=2).value
    credit_type = sheet.cell(row=i, column=4).value
    overdraft_type = sheet.cell(row=i, column=5).value
    credit_limit_type = sheet.cell(row=i, column=9).value
    LLP_group =  sheet.cell(row=i, column=14).value
    LLP_ratio =  sheet.cell(row=i, column=15).value

   #выбираем клиентов с кредитами, овердрафтами, кредитн.линиями и имеющими категорию качества ниже первой.
   # Строим четыре хеша с данными по каждой группе риска
    if LLP_group == required_LLP_group and (credit_type == point_in_xlsx or overdraft_type == point_in_xlsx or credit_limit_type == point_in_xlsx):
      #print(customer_id)
      if customer_id is None or customer_id in added_customer_id:
        continue
    #print(bank_id, bank_name, LLP_group, LLP_ratio)
      res.append({
        'LLP_group': LLP_group,
        'customer_id': customer_id, # ternar operator for
        'customer_name': customer_name,
        'LLP_ratio': round(LLP_ratio*100)
        #'credit_type': credit_type,
        #'overdraft_type': overdraft_type,
        #'credit_limit_type': credit_limit_type
      })
      added_customer_id.update({customer_id: True})
  return res



def list_rows():
  #склеиваем 4 хеша созданные в функции выше в один
  risk_loans = np.concatenate((two_risk_group, three_risk_group, four_risk_group, five_risk_group), axis=0)
  #print(risk_loans)
  #print(risk_loans[0]['customer_id'])
  return risk_loans


def f115_portf(file_path):
  book = openpyxl.load_workbook(file_path)
  sheet = book['clients']
  sheet2 = book['pivot']
  debt_dict = {}
  total_debt_dict = {}
  # full_id_list = []
  for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row=row, column=1).value
    #восстанавливаем полный id клиента
    full_id = str('956')+str(cell)
    debt = sheet.cell(row=row, column=5).value
    #строим таблицу: клиент:размер задолженности
    if full_id in debt_dict:
     debt_dict[full_id] = debt_dict.get(full_id) + debt/1000 #если несколько сумм у одного клиента, они складываются
    else:
      debt_dict.update({full_id : debt/1000})
  total_debt_dict = {
  2:sheet2.cell(row=9, column=3).value/1000,
  3:sheet2.cell(row=10, column=3).value/1000,
  4:sheet2.cell(row=11, column=3).value/1000,
  5:sheet2.cell(row=12, column=3).value/1000
  }
  #print (total_debt_dict)
  return debt_dict, total_debt_dict


def comis_proc(file_path):

  book = openpyxl.load_workbook(file_path)
  sheet = book['f115']
  sheet2 = book['Pivot']
  proc_dict = {}
  comis_dict = {}
  total_proc_dict = {}
  total_comis_dict = {}
  total_proc_llp_dict = {}
  total_comis_llp_dict = {}
  # full_id_list = []
  for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row=row, column=7).value
    full_id = str('956')+str(cell)
    claim_type = sheet.cell(row=row, column=3).value
    claim_detail = sheet.cell(row=row, column=4).value
    claim_amount = sheet.cell(row=row, column=8).value
    reserve_claim_amount = nv.check_value(sheet.cell(row=row, column=9).value)
    if claim_type == 'Проценты':
      if full_id in proc_dict:
        #строим хеш клиент: массив(сумма процентов, сумма резерва под проценты), аналогично по комиссиям
        proc_dict[full_id][0] = proc_dict.get(full_id)[0] + claim_amount/1000
        proc_dict[full_id][1] = proc_dict.get(full_id)[1] + reserve_claim_amount/1000
      else:
        proc_dict.update({full_id : [claim_amount/1000, reserve_claim_amount/1000]})
    # если комиссия содержит текст "за резервирование"
    elif claim_type == 'Комиссии' and re.search(r'за\s+резервирование\s+', claim_detail):
      if full_id in comis_dict:
        comis_dict[full_id][0] = comis_dict.get(full_id)[0] + claim_amount/1000
        comis_dict[full_id][1] = comis_dict.get(full_id)[1] + reserve_claim_amount/1000
      else:
        comis_dict.update({full_id : [claim_amount/1000, reserve_claim_amount/1000]})

  # val = driver_excel_cell.check_empty_number_cell(sheet2.cell(row=9, column=4).value)

  for row2 in range(8, 13):
    for column2 in range(4, 9):
       sheet2.cell(row=row2, column=column2).value = nv.check_value(sheet2.cell(row=row2, column=column2).value)

  total_comis_dict = {
    2:[sheet2.cell(row=9, column=4).value/1000 , sheet2.cell(row=9, column=5).value/1000],
    3:[sheet2.cell(row=10, column=4).value/1000, sheet2.cell(row=10, column=5).value/1000],
    4:[sheet2.cell(row=11, column=4).value/1000, sheet2.cell(row=11, column=5).value/1000],
    5:[sheet2.cell(row=12, column=4).value/1000, sheet2.cell(row=12, column=5).value/1000]
  }

  total_proc_dict = {
    2:[sheet2.cell(row=9, column=6).value/1000, sheet2.cell(row=9, column=7).value/1000],
    3:[sheet2.cell(row=10, column=6).value/1000, sheet2.cell(row=10, column=7).value/1000],
    4:[sheet2.cell(row=11, column=6).value/1000, sheet2.cell(row=11, column=7).value/1000],
    5:[sheet2.cell(row=12, column=6).value/1000, sheet2.cell(row=12, column=7).value/1000]
  }

  #print(total_comis_dict)
  #print(total_proc_dict)
  #print(proc_dict)
  #print(comis_dict)
  return proc_dict, comis_dict, total_comis_dict, total_proc_dict


def reserve_principal(file_path):
  book = openpyxl.load_workbook(file_path)
  sheet = book['Rez-Bal']
  reserve_dict = {}
  total_reserve_dict = {}
  sheet2 = book['pivot']
  #строим хеш: клиент: сумма резерва по основному долгу

  for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row=row, column=1).value
    full_id = str('956')+str(cell)
    reserve = nv.check_value(sheet.cell(row=row, column=5).value)
    if full_id in reserve_dict:
      reserve_dict[full_id] = reserve_dict.get(full_id) + reserve/1000
    else:
      reserve_dict.update({full_id : reserve/1000})
  #собираем общее из таблицы pivot
  total_reserve_dict = {
  2:sheet2.cell(row=5, column=4).value/1000,
  3:sheet2.cell(row=7, column=4).value/1000,
  4:sheet2.cell(row=9, column=4).value/1000,
  5:sheet2.cell(row=11, column=4).value/1000
  }
  #print (total_reserve_dict)
  #print(reserve_dict)
  return reserve_dict, total_reserve_dict



def write_xlsx(list_rows, debt_dict, proc_dict, comis_dict, reserve_dict, total_debt_dict, total_proc_dict, total_comis_dict, total_reserve_dict):
  #запись в файл
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  ws.title = 'data'
  #ws = wb.create_sheet(title='data')
  ws2 = wb.create_sheet(title='total')
  ws.column_dimensions['B'].width = 15
  ws.column_dimensions['C'].width = 30

  column = 4
  while column < 11:
    x = get_column_letter(column)
    ws.column_dimensions[x].width = 15
    column += 1

  column = 2
  while column < 11:
    x = get_column_letter(column)
    ws2.column_dimensions[x].width = 15
    column += 1

  #задаем заголовки столбцов
  ws.cell(row=1, column=1).value = 'LLP_group'
  ws.cell(row=1, column=2).value = 'customer_id'
  ws.cell(row=1, column=3).value = 'customer_name'
  ws.cell(row=1, column=4).value = 'principal_amount'
  ws.cell(row=1, column=5).value = 'interest_amount'
  ws.cell(row=1, column=6).value = 'commission_amount'
  ws.cell(row=1, column=7).value ='LLP_ratio'
  ws.cell(row=1, column=8).value ='LLP_potential'
  ws.cell(row=1, column=9).value ='LLP_real'
  ws.cell(row=1, column=10).value ='LLP_interest_real'
  ws.cell(row=1, column=11).value ='LLP_commission_real'

  for i in range(0, len(list_rows)):
      full_id = list_rows[i]['customer_id']
      row = i+2

      ws.cell(row=row, column=1).value = list_rows[i]['LLP_group']
      short_id = str(full_id)[3:]
      ws.cell(row=row, column=2).value = int(short_id)
      ws.cell(row=row, column=3).value = list_rows[i]['customer_name']
      ws.cell(row=row, column=4).value = debt_dict.get(str(full_id), 0)
      ws.cell(row=row, column=4).number_format = '# ### ##0.000'

      ws.cell(row=row, column=5).value = proc_dict.get(str(full_id), [0,0])[0]
      ws.cell(row=row, column=5).number_format = '# ### ##0.000'

      ws.cell(row=row, column=6).value = comis_dict.get(str(full_id), [0,0])[0]
      ws.cell(row=row, column=6).number_format = '# ### ##0.000'

      ws.cell(row=row, column=7).value = list_rows[i]['LLP_ratio']
      ws.cell(row=row, column=8).value = ws.cell(row=row, column=4).value * ws.cell(row=row, column=7).value/100
      ws.cell(row=row, column=8).number_format = '# ### ##0.000'
      ws.cell(row=row, column=9).value = reserve_dict.get(str(full_id), 0)
      ws.cell(row=row, column=9).number_format = '# ### ##0.000'
      ws.cell(row=row, column=10).value = proc_dict.get(str(full_id), [0,0])[1]
      ws.cell(row=row, column=10).number_format = '# ### ##0.000'
      ws.cell(row=row, column = 11).value = comis_dict.get(str(full_id), [0,0])[1]
      ws.cell(row=row, column=11).number_format = '# ### ##0.000'



  #задаем заголовки столбцов
  ws2.cell(row=1, column=1).value = 'LLP_group'
  ws2.cell(row=1, column=2).value = 'principal_amount total'
  ws2.cell(row=1, column=3).value = 'interest_amount total'
  ws2.cell(row=1, column=4).value = 'commission_amount total'
  ws2.cell(row=1, column=5).value = 'LLP_potential total'
  ws2.cell(row=1, column=6).value = 'LLP_real total'
  ws2.cell(row=1, column=7).value = 'LLP_interest_real total'
  ws2.cell(row=1, column=8).value = 'LLP_commission_real total'


  for i in range(2,6):
      row2 = i
      ws2.cell(row=row2, column=1).value = i
      ws2.cell(row=row2, column=2).value = total_debt_dict[i]
      ws2.cell(row=row2, column=2).number_format = '# ### ##0.000'
      ws2.cell(row=row2, column=3).value = total_proc_dict[i][0]
      ws2.cell(row=row2, column=3).number_format = '# ### ##0.000'
      ws2.cell(row=row2, column=4).value = total_comis_dict[i][0]
      ws2.cell(row=row2, column=4).number_format = '# ### ##0.000'
      ws2.cell(row=row2, column=6).value = total_reserve_dict[i]
      ws2.cell(row=row2, column=6).number_format = '# ### ##0.000'
      ws2.cell(row=row2, column=7).value = total_proc_dict[i][1]
      ws2.cell(row=row2, column=7).number_format = '# ### ##0.000'
      ws2.cell(row=row2, column=8).value = total_comis_dict[i][1]
      ws2.cell(row=row2, column=8).number_format = '### ##0.000'
      ws2.cell(row=row2, column=9).value = 'peschkova pivot'
  wb.save(filename='../doc_form_115/data_115.xlsx')

def get_sheet_name():
  current_date = date.today()
  this_month = current_date.strftime('%m')
  this_month_number = int(this_month)
  if this_month_number > 1:
    last_month_number = this_month_number - 1
    year_number = int(current_date.strftime('%Y'))
  elif this_month_number == 1:
    last_month_number = 12
    year_number = int(current_date.strftime('%Y')) - 1
  last_month_date = date(year_number, last_month_number, 1)
  sheet_name =  last_month_date.strftime('%B_%Y')
  # print(sheet_name)
  return(sheet_name)


if __name__ == "__main__":
  book = openpyxl.load_workbook('../doc_form_115/LLP.xlsx')
  sheet = book[get_sheet_name()]

  # TODO: вместо файла читать вкладку
  # list_row = read_xlsx('./LLP.xlsx')
  two_risk_group = filter_by_LLP_group(sheet, 2)
  three_risk_group = filter_by_LLP_group(sheet, 3)
  four_risk_group = filter_by_LLP_group(sheet, 4)
  five_risk_group = filter_by_LLP_group(sheet, 5)

  proc_dict, comis_dict, total_comis_dict, total_proc_dict = comis_proc('../doc_form_115/RVP_Comis_Proc.xlsx')
  reserve_dict, total_reserve_dict = reserve_principal('../doc_form_115/Rez_Bal.xlsx')
  debt_dict, total_debt_dict = f115_portf('../doc_form_115/f115_portf.xlsx')


  # f115_portf('./f115_portf.xlsx', list_rows())
  write_xlsx(list_rows(), debt_dict, proc_dict, comis_dict, reserve_dict, total_debt_dict, total_proc_dict, total_comis_dict, total_reserve_dict)
