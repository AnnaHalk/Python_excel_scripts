#!/usr/local/bin/python3

import openpyxl
import numpy as np
import re


def sum_data(sheet, required_LLP_group):
  res = []
  principal_amount_total = 0
  interest_amount_total = 0
  commission_amount_total = 0
  LLP_potential_total = 0
  LLP_real_total = 0
  LLP_interest_real_total = 0
  LLP_commission_real_total = 0

  for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == required_LLP_group:
      if sheet.cell(row=row, column=2).value is not 0:
        principal_amount_total += sheet.cell(row=row, column=4).value
        interest_amount_total += sheet.cell(row=row, column=5).value
        commission_amount_total += sheet.cell(row=row, column=6).value
        LLP_potential_total += sheet.cell(row=row, column=8).value
        LLP_real_total += sheet.cell(row=row, column=9).value
        LLP_interest_real_total += sheet.cell(row=row, column=10).value
        LLP_commission_real_total += sheet.cell(row=row, column=11).value

  res.append({
    'LLP_group': required_LLP_group,
    'principal_amount total': principal_amount_total,
    'interest_amount total': interest_amount_total,
    'commission_amount total': commission_amount_total,
    'LLP_potential total': LLP_potential_total,
    'LLP_real total': LLP_real_total,
    'LLP_interest_real total': LLP_interest_real_total,
    'LLP_commission_real total': LLP_commission_real_total,
  })
  return res


def list_sums():
  # склеиваем 4 хеша созданные в функции выше в один
  sum_data = np.concatenate(
      (two_risk_group, three_risk_group, four_risk_group, five_risk_group), axis=0)
  # print(sum_data)
  return sum_data


def write_xlsx(list_sums):
  # запись в файл
  wb = openpyxl.load_workbook('../doc_form_115/./data_115.xlsx')
  ws = wb['total']

  for i in range(0, len(list_sums)):
    row = i + 6
    ws.cell(row=row, column=1).value = list_sums[i]['LLP_group']
    ws.cell(row=row, column=2).value = list_sums[i]['principal_amount total']
    ws.cell(row=row, column=3).value = list_sums[i]['interest_amount total']
    ws.cell(row=row, column=4).value = list_sums[i]['commission_amount total']
    ws.cell(row=row, column=5).value = list_sums[i]['LLP_potential total']
    ws.cell(row=row, column=6).value = list_sums[i]['LLP_real total']
    ws.cell(row=row, column=7).value = list_sums[i]['LLP_interest_real total']
    ws.cell(
        row=row, column=8).value = list_sums[i]['LLP_commission_real total']
    ws.cell(row=row, column=9).value = 'calculated'

  ws.cell(row=10, column=1).value = 'delta 2 group'
  ws.cell(row=11, column=1).value = 'delta 3 group'
  ws.cell(row=12, column=1).value = 'delta 4 group'
  ws.cell(row=13, column=1).value = 'delta 5 group'

  for y in range(2, 5):
    ws.cell(row=10, column=y).value = ws.cell(
        row=2, column=y).value - ws.cell(row=6, column=y).value
    ws.cell(row=11, column=y).value = ws.cell(
        row=3, column=y).value - ws.cell(row=7, column=y).value
    ws.cell(row=12, column=y).value = ws.cell(
        row=4, column=y).value - ws.cell(row=8, column=y).value
    ws.cell(row=13, column=y).value = ws.cell(
        row=5, column=y).value - ws.cell(row=9, column=y).value

  for y in range(6, 9):
    ws.cell(row=10, column=y).value = ws.cell(
        row=2, column=y).value - ws.cell(row=6, column=y).value
    ws.cell(row=11, column=y).value = ws.cell(
        row=3, column=y).value - ws.cell(row=7, column=y).value
    ws.cell(row=12, column=y).value = ws.cell(
        row=4, column=y).value - ws.cell(row=8, column=y).value
    ws.cell(row=13, column=y).value = ws.cell(
        row=5, column=y).value - ws.cell(row=9, column=y).value

  wb.save(filename='data_115.xlsx')


if __name__ == "__main__":

  book = openpyxl.load_workbook('../doc_form_115/./data_115.xlsx')
  sheet = book['data']

  two_risk_group = sum_data(sheet, 2)
  three_risk_group = sum_data(sheet, 3)
  four_risk_group = sum_data(sheet, 4)
  five_risk_group = sum_data(sheet, 5)

  list_sums()


write_xlsx(list_sums())
